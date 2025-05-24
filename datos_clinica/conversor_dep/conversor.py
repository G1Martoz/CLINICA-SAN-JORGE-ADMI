import os
import sys
import tkinter as tk
from tkinter import messagebox
import xlrd
from openpyxl import Workbook
from datetime import datetime
import shutil

# --- CONFIGURACIÓN ---
ARCHIVOS_OBJETIVO = [
    "disponibilidad.xls",
    "distribucion.xls",
    "errores.xls",
    "nomina.xls",
    "produccion.xls"
]
CARPETA_SALIDA = "convertidos"
NOMBRE_HOJA_UNIFICADA = "Hoja 1"

# --- FUNCIÓN PRINCIPAL ---
def convertir_automaticamente():
    """
    Busca archivos .xls, los convierte a .xlsx. Antes de copiar cada hoja,
    calcula la última fila con datos reales (basado en contenido en la primera columna)
    para ignorar la "basura" final y limpia la carpeta de destino antes de cada ejecución.
    """
    script_dir = os.path.dirname(sys.executable)
    ruta_salida = os.path.join(script_dir, CARPETA_SALIDA)

    if os.path.exists(ruta_salida):
        shutil.rmtree(ruta_salida)
    os.makedirs(ruta_salida)

    archivos_encontrados = []
    for archivo in os.listdir(script_dir):
        if archivo.lower() in ARCHIVOS_OBJETIVO:
            archivos_encontrados.append(os.path.join(script_dir, archivo))

    if not archivos_encontrados:
        messagebox.showwarning("No se encontraron archivos", "No se encontró ninguno de los archivos objetivo en la carpeta.")
        return

    contador_convertidos = 0
    for ruta_archivo in archivos_encontrados:
        try:
            libro_xls = xlrd.open_workbook(ruta_archivo, formatting_info=False)
            libro_xlsx = Workbook()
            libro_xlsx.remove(libro_xlsx.active)
            hoja_destino = libro_xlsx.create_sheet(title=NOMBRE_HOJA_UNIFICADA)

            for indice_hoja in range(libro_xls.nsheets):
                hoja_xls = libro_xls.sheet_by_index(indice_hoja)
                
                last_real_row = -1
                for row_num in range(hoja_xls.nrows - 1, -1, -1):
                    row_values = hoja_xls.row_values(row_num)
                    
                    # --- *** EL CAMBIO DEFINITIVO ESTÁ AQUÍ *** ---
                    # Antes: if any(str(v).strip() for v in row_values):
                    # Ahora: Una fila es válida solo si tiene algo en la primera columna.
                    if row_values and str(row_values[0]).strip():
                        last_real_row = row_num
                        break
                
                if last_real_row == -1:
                    continue

                fila_inicio = 1 if indice_hoja > 0 else 0

                for fila_idx in range(fila_inicio, last_real_row + 1):
                    fila_valores = []
                    for col_idx in range(hoja_xls.ncols):
                        celda = hoja_xls.cell(fila_idx, col_idx)
                        valor = celda.value
                        if celda.ctype == xlrd.XL_CELL_DATE:
                            try:
                                valor = datetime(*xlrd.xldate_as_tuple(valor, libro_xls.datemode))
                            except ValueError:
                                pass
                        fila_valores.append(valor)
                    
                    hoja_destino.append(fila_valores)

            nombre_base = os.path.basename(ruta_archivo)
            nombre_nuevo = os.path.splitext(nombre_base)[0] + ".xlsx"
            ruta_guardado = os.path.join(ruta_salida, nombre_nuevo)
            libro_xlsx.save(ruta_guardado)
            
            contador_convertidos += 1
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al convertir el archivo:\n{os.path.basename(ruta_archivo)}\n\nError: {e}")
            return

    messagebox.showinfo(
        "Conversión Finalizada",
        f"Proceso completado con éxito.\n\nSe convirtieron {contador_convertidos} archivos.\n\nGuardados en la carpeta: '{CARPETA_SALIDA}'"
    )

# --- BLOQUE DE EJECUCIÓN PRINCIPAL ---
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    convertir_automaticamente()